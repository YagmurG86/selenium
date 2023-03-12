from openpyxl import Workbook, load_workbook

# =======================================

# wb = Workbook()
# ws = wb.active
# ws["A1"] = "Yagmur"
# wb.save("PythonForTesters\demo_write.xlsx")

# =======================================

# wb = Workbook()
# ws = wb.active
# testdate = [["Name", "City"], ["Yagmur", "Turkey"], ["Elvis", "Bosnia"]]

# for data in testdate:
#     ws.append(data)

# wb.save("PythonForTesters\demo_write.xlsx")

# =======================================

# wb = Workbook()
# ws = wb.active

# for i in range(1, 6):
#     for j in range(1, 5):
#         ws.cell(row=i, column=j).value = i+j

# wb.save("PythonForTesters\demo_write.xlsx")

# =======================================

wb = load_workbook(filename="PythonForTesters\demo_read.xlsx")
# ws = wb.active
ws = wb["Sheet"]

# print(ws["A3"].value)
# print(wb["DemoSheet"]["A2"].value)
# print(ws.cell(2, 2).value)

for i in range(1, ws.max_row+1):
    for j in range(1, ws.max_column+1):
        print(ws.cell(row=i, column=j).value)
